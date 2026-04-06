import argparse
import datetime as dt
import json
import os
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def _jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def _is_row_empty(values: List[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _find_first_non_empty_row(ws, max_scan_rows: int) -> Optional[int]:
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        if not _is_row_empty(list(row)):
            return row_idx
    return None


def _load_sheet(
    wb,
    sheet_name: str,
    header_row: Optional[int],
    data_start_row: Optional[int],
    max_scan_rows: int,
    max_rows: Optional[int],
) -> Dict[str, Any]:
    ws = wb[sheet_name]

    inferred_header_row = header_row
    if inferred_header_row is None:
        inferred_header_row = _find_first_non_empty_row(ws, max_scan_rows=max_scan_rows)
    if inferred_header_row is None:
        return {"headers": [], "rows": []}

    header_values = list(
        ws.iter_rows(
            min_row=inferred_header_row,
            max_row=inferred_header_row,
            values_only=True,
        )
    )[0]
    headers: List[str] = []
    used: Dict[str, int] = {}
    for idx, raw in enumerate(header_values, start=1):
        base = str(raw).strip() if raw is not None and str(raw).strip() else f"col_{idx}"
        n = used.get(base, 0) + 1
        used[base] = n
        headers.append(base if n == 1 else f"{base}__{n}")

    start_row = data_start_row if data_start_row is not None else inferred_header_row + 1

    rows: List[Dict[str, Any]] = []
    taken = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        values = list(row)
        if _is_row_empty(values):
            continue
        record = {headers[i]: _jsonable(values[i]) if i < len(values) else None for i in range(len(headers))}
        rows.append(record)
        taken += 1
        if max_rows is not None and taken >= max_rows:
            break

    return {"headers": headers, "rows": rows}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Export an Excel workbook (.xlsx) to JSON for KPI analysis."
    )
    p.add_argument(
        "--input",
        "-i",
        default=r"C:\Users\Jake\OneDrive - Wilde Signs\Production List\Archive\Archive.xlsx",
        help="Path to .xlsx workbook.",
    )
    p.add_argument(
        "--output",
        "-o",
        default=None,
        help="Path to output JSON file. Defaults to <input>.json",
    )
    p.add_argument(
        "--sheets",
        default=None,
        help="Comma-separated list of sheet names to export (default: all).",
    )
    p.add_argument(
        "--header-row",
        type=int,
        default=None,
        help="1-based row index to treat as header row (default: inferred).",
    )
    p.add_argument(
        "--data-start-row",
        type=int,
        default=None,
        help="1-based row index to start reading data rows (default: header_row + 1).",
    )
    p.add_argument(
        "--max-scan-rows",
        type=int,
        default=50,
        help="Max rows to scan to infer header row (default: 50).",
    )
    p.add_argument(
        "--max-rows",
        type=int,
        default=None,
        help="Max non-empty data rows to export per sheet.",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    input_path = os.path.abspath(args.input)
    output_path = (
        os.path.abspath(args.output)
        if args.output
        else os.path.splitext(input_path)[0] + ".json"
    )

    wb = openpyxl.load_workbook(
        input_path,
        data_only=True,
        read_only=True,
    )
    try:
        wanted_sheets: List[str]
        if args.sheets:
            wanted_sheets = [s.strip() for s in args.sheets.split(",") if s.strip()]
        else:
            wanted_sheets = list(wb.sheetnames)

        missing = [s for s in wanted_sheets if s not in wb.sheetnames]
        if missing:
            raise SystemExit(f"Missing sheet(s): {', '.join(missing)}")

        payload: Dict[str, Any] = {
            "source": input_path,
            "exportedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
            "sheets": {},
        }

        for sheet_name in wanted_sheets:
            payload["sheets"][sheet_name] = _load_sheet(
                wb,
                sheet_name=sheet_name,
                header_row=args.header_row,
                data_start_row=args.data_start_row,
                max_scan_rows=args.max_scan_rows,
                max_rows=args.max_rows,
            )

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        print(f"Wrote: {output_path}")
        for sheet_name in wanted_sheets:
            sheet = payload["sheets"][sheet_name]
            print(f"  {sheet_name}: {len(sheet['rows'])} rows, {len(sheet['headers'])} columns")
        return 0
    finally:
        wb.close()


if __name__ == "__main__":
    raise SystemExit(main())
