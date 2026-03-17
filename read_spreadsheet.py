import openpyxl
import json

wb = openpyxl.load_workbook(
    r'C:\Users\Jake\OneDrive - Wilde Signs\Production List\February 2026\Production List_02_10_26.xlsm',
    data_only=True,
    read_only=True
)

ws = wb['Sheet1']
print('=== Full Sheet1 data (rows 50-150) ===')
for i, row in enumerate(ws.iter_rows(min_row=50, max_row=150, values_only=False)):
    vals = []
    for c in row:
        if c.value is not None:
            vals.append((str(c.value)[:80], c.coordinate))
    if vals:
        print(f'  Row {i+50}: {vals}')

print()
print('=== Unique values in column J (routing) ===')
j_vals = set()
for row in wb['Sheet1'].iter_rows(min_row=4, max_row=200, min_col=10, max_col=10, values_only=True):
    if row[0] and str(row[0]).strip():
        j_vals.add(str(row[0]).strip())
print(sorted(j_vals))

print()
print('=== Unique values in column B (WO #) - checking format ===')
b_vals = []
for row in wb['Sheet1'].iter_rows(min_row=4, max_row=200, min_col=2, max_col=2, values_only=True):
    if row[0] and str(row[0]).strip() and str(row[0]).strip() != '-':
        b_vals.append(str(row[0]).strip())
print(f'Total WO entries: {len(b_vals)}')
print(f'All: {b_vals}')

print()
print('=== Total row count ===')
total = 0
for row in wb['Sheet1'].iter_rows(max_row=500, values_only=True):
    if any(v is not None for v in row):
        total += 1
print(f'Total non-empty rows: {total}')

wb.close()
